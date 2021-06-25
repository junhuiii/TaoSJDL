import time
import glob
import re
import openpyxl
import xlrd
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.wait import WebDriverWait
from openpyxl.utils import exceptions

# Defined variables
login_payload = {'phone_num': '91784364', 'pw': 'fupin123'}
login_url = 'https://login.taosj.com/?redirectURL=https%3A%2F%2Fwww.taosj.com%2F'
taosj_meta_data = r'C:\Users\Dell\IdeaProjects\TaoSJDL\src\TaoSJ Meta'

# Arguments for chrome webdriver
option = webdriver.ChromeOptions()
option.add_argument('--disable-notifications')
option.add_argument("--mute-audio")
option.add_argument("""user-agent= Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) 
                        Chrome/58.0.3029.110 Safari/537.36""")
option.add_argument("--incognito")
option.add_experimental_option('excludeSwitches', ['enable-automation'])
option.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})


def read_xlsx_file(xlsx_file_path):
    wb = openpyxl.load_workbook(xlsx_file_path)
    ws = wb.active
    brand = ws['C2'].value
    category = ws['D2'].value
    return brand, category


def read_xls_file(xls_file_path):
    wb = xlrd.open_workbook(xls_file_path)
    ws = wb.sheet_by_index(0)
    brand = ws.cell_value(rowx=1, colx=2)
    category = ws.cell_value(rowx=1, colx=3)
    return brand, category


def click_xpath(html_xpath):
    element = driver.find_element_by_xpath(html_xpath)
    element.click()


def send_keys(html_xpath, key):
    element = driver.find_element_by_xpath(html_xpath)
    element.send_keys(key)


def login_process():
    try:
        # Select country code
        country_xpath = '//*[@id="J_Mod_Login"]/form/div[2]/span'
        click_xpath(country_xpath)

        singapore_xpath = '//*[@id="J_Mod_Login"]/form/div[2]/div/p[6]'
        click_xpath(singapore_xpath)

        time.sleep(3)

        # Input phone num and pw
        phone_num = '//*[@id="J_Mod_Login"]/form/div[2]/input'
        send_keys(phone_num, login_payload['phone_num'])

        passw = '//*[@id="J_Mod_Login"]/form/div[3]/input'
        send_keys(passw, login_payload['pw'])

        time.sleep(3)

        # Click on login button
        login = '//*[@id="T_Login"]'
        click_xpath(login)
        print("Log In Successful.")

    except Exception as e:
        print(e)


def mouse_over(xpath):
    element = driver.find_element_by_xpath(xpath)
    hover = ActionChains(driver).move_to_element(element)
    hover.perform()


def shop_data():
    try:
        # Navigate to '找宝贝' Tab
        china_services = '//*[@id="J_c-data-top"]/div/div[2]/ul/li[2]/div[2]'  # '国内电商‘ Element
        mouse_over(china_services)

        shop_data = '//*[@id="J_c-data-top"]/div/div[2]/ul/li[2]/div[3]/div/div[1]/a[3]'  # '店铺数据’ Element
        click_xpath(shop_data)

        time.sleep(5)

        driver.switch_to.window(driver.window_handles[-1])
        sku_lookup = '//*[@id="header"]/div/div[6]/div/div[1]/div/ul/li[2]/a'
        click_xpath(sku_lookup)
        print("Successfully Navigated to '找宝贝‘ Tab. ")

    except Exception as e:
        print(e)


if __name__ == '__main__':

    # Obtain list of files containing SKU IDs to scrape from directory
    list_of_xlsx = glob.glob('**/TaoSJ Meta/*.xlsx', recursive=True) + \
                   glob.glob('**/TaoSJ Meta/*.xls', recursive=True)
    len_list_of_xlsx = len(list_of_xlsx)
    print(f"Found {len_list_of_xlsx} SKUs to scrape.")

    # Obtain data regarding each SKU, outputs destination path(str) for downloading
    sku_id_regex = re.compile('\\\\([0-9]*).xls')
    list_sku_id = []
    brands = []
    categories = []
    for filepath in list_of_xlsx:

        sku_info = {}
        # Get SKU_ID and put inside sku_info dictionary
        sku_id = sku_id_regex.search(filepath).group(1)
        list_sku_id.append(sku_id)
        sku_info['sku_id'] = sku_id

        # Get brand and put inside sku_info_dictionary
        try:
            brand, category = read_xlsx_file(filepath)
            brands.append(brand)
            categories.append(category)
            sku_info['brand'] = brand
            sku_info['category'] = category

        except Exception:
            brand,category = read_xls_file(filepath)
            brands.append(brand)
            categories.append(category)
            sku_info['brand'] = brand
            sku_info['category'] = category
        #TODO: Clean up directory format, test on tf-es-dumpling script
        #TODO: Based on brand and category, create a path to save the file to for a particular sku_id

    #Login Process
    driver = webdriver.Chrome('chromedriveedit.exe', options=option)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})""",
    })
    driver.get(login_url)
    wait = WebDriverWait(driver,10)
    login_process()

    time.sleep(5)

    #Navigate to '找宝贝' Tab
    shop_data()
