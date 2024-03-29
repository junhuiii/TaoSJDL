import re

import openpyxl
import pytoml
import xlrd
from openpyxl.utils.exceptions import InvalidFileException
from selenium import webdriver
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

from file_downloading import *
from move_files import *
from sort_file_path import *

# config path
CONFIG_PATH = 'config.toml'
PROJECT_PATH = os.getcwd()

# Arguments for chrome webdriver
option = webdriver.ChromeOptions()
option.add_argument('--disable-notifications')
option.add_argument("--mute-audio")
option.add_argument("""user-agent= Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) 
                        Chrome/58.0.3029.110 Safari/537.36""")
option.add_argument("--incognito")
option.add_experimental_option('excludeSwitches', ['enable-automation'])
option.add_experimental_option("prefs", {"profile.managed_default_content_settings.images": 2})
prefs = {'download.default_directory': PROJECT_PATH + '\download-dump'}
option.add_experimental_option('prefs', prefs)
option.add_argument("enable-features=NetworkServiceInProcess")
option.add_argument("--window-size=1920,1080")


# Read config.toml file
def read_config(path):
    config = pytoml.load(open(path, 'rb'))
    return config


def read_xlsx_file(xlsx_file_path):
    wb = openpyxl.load_workbook(xlsx_file_path)
    ws = wb.active
    brand = ws['C2'].value
    category = ws['D2'].value
    return brand, category


def read_xls_file(xls_file_path):
    wb = xlrd.open_workbook(xls_file_path)
    ws = wb.sheet_by_index(0)
    brand = ws.cell_value(rowx=1, colx=2)
    category = ws.cell_value(rowx=1, colx=3)
    return brand, category


def click_xpath(html_xpath):
    element = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, html_xpath)))
    element.click()


def send_keys(html_xpath, key):
    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, html_xpath)))
    element.send_keys(key)


def clear_text_field(html_xpath):
    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, html_xpath)))
    element.clear()


def login_process():
    try:
        # Select country code
        country_xpath = '//*[@id="J_Mod_Login"]/form/div[2]/span'
        click_xpath(country_xpath)

        singapore_xpath = '//*[@id="J_Mod_Login"]/form/div[2]/div/p[6]'
        click_xpath(singapore_xpath)

        # Input phone num and pw
        phone_num = '//*[@id="J_Mod_Login"]/form/div[2]/input'
        send_keys(phone_num, config_file['defined_vars']['taosj_phone_num'])

        passw = '//*[@id="J_Mod_Login"]/form/div[3]/input'
        send_keys(passw, config_file['defined_vars']['taosj_pw'])

        time.sleep(1)

        # Click on login button
        login = '//*[@id="T_Login"]'
        click_xpath(login)
        print("Log In Successful.")

    except Exception as login_error:
        print(login_error)


def mouse_over(xpath):
    element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, xpath)))
    hover = ActionChains(driver).move_to_element(element)
    hover.perform()


def shop_data():
    try:
        # Navigate to '找宝贝' Tab
        china_services = '//*[@id="J_c-data-top"]/div/div[2]/ul/li[2]/div[2]'  # '国内电商‘ Element
        mouse_over(china_services)

        shop_data = '//*[@id="J_c-data-top"]/div/div[2]/ul/li[2]/div[3]/div/div[1]/a[3]'  # '店铺数据’ Element
        click_xpath(shop_data)

        time.sleep(8)

        driver.switch_to.window(driver.window_handles[-1])
        sku_lookup = '//*[@id="header"]/div/div[6]/div/div[1]/div/ul/li[2]/a'
        time.sleep(3)
        click_xpath(sku_lookup)
        print("Successfully Navigated to '找宝贝‘ Tab. ")

    except Exception as e:
        print(e)


def scrape_sku(sku_id):
    time.sleep(2)
    # Input sku_id into search bar
    sku_input_field_xpath = '//*[@id="search"]'
    send_keys(sku_input_field_xpath, sku_id)
    time.sleep(5)
    # Click on search button
    sku_search_button = '//*[@id="header"]/div/div[6]/div/div[1]/div/div/a'
    click_xpath(sku_search_button)

    time.sleep(2)

    # Look for SKU in search results
    # Error handling: Can either be found or not be found

    try:
        sku_exists = driver.find_element_by_xpath(f'//a[contains(@href,"//item.taobao.com/item.htm?id={sku_id}")]')
        sku_data_button = driver.find_element_by_xpath(
            '//*[@id="container"]/div/div[4]/div/div/div[2]/table/tbody/tr/td[7]/div/div/a')
        print(f"{sku_id} exists. Found data on it.")
        download_data(sku_data_button, sku_id)
        driver.switch_to.window(driver.window_handles[-1])
        clear_text_field(sku_input_field_xpath)

    except NoSuchElementException:

        try:
            sku_no_exist = driver.find_element_by_xpath('//*[@id="container"]/div/div[4]/div/div/div[1]/form/div/span[1]')
            print(f"{sku_id} does not exist.")
            clear_text_field(sku_input_field_xpath)

        except NoSuchElementException:
            sku_no_exist = driver.find_element_by_xpath('//div[text()="暂无数据"]')
            #sku_no_exist = driver.find_element_by_xpath('//*[@id="container"]/div/div[3]/div/div/form/div/p')
            print(f"{sku_id} does not exist.")
            clear_text_field(sku_input_field_xpath)


# TODO: Deal with long waiting time because of hanging, implement refresh function
def download_data(selenium_element, sku_id):
    selenium_element.click()
    time.sleep(3)
    driver.switch_to.window(driver.window_handles[-1])

    driver.implicitly_wait(30)
    time.sleep(5)
    download_button = '//*[@id="itemMain"]/div/div[4]/div/div[1]/a'
    retries = 1
    while retries <=5:
        try:
            print(f"Downloading data for {sku_id}.....")
            waiter = FileWaiter(PROJECT_PATH + config_file['defined_vars']['download_dump_folder'] + '\\*.xls')
            click_xpath(download_button)
            result_of_download = waiter.wait_new_file(15)

            while result_of_download == 'File did not download.':
                print(f'{result_of_download}, trying again...')
                click_xpath(download_button)
                result_of_download = waiter.wait_new_file(15)
                retries += 1
                if result_of_download != 'File did not download.' or retries == 5:
                    break
            print(result_of_download)
            break

        except TimeoutException as timeout:
            driver.refresh()
            print(f'{timeout} occurred. Will try again...')
            retries += 1
            pass

    time.sleep(10)
    driver.close()


def wait_for_downloads(waiter):
    file_downloading = waiter.wait_new_file(10)
    return file_downloading


if __name__ == '__main__':

    # Obtain list of files containing SKU IDs to scrape from directory
    list_of_xlsx = glob.glob('**/TaoSJ Meta/*.xlsx', recursive=True) + \
                   glob.glob('**/TaoSJ Meta/*.xls', recursive=True)
    len_list_of_xlsx = len(list_of_xlsx)
    print(f"Found {len_list_of_xlsx} SKUs to scrape.")

    # Obtain data regarding each SKU, outputs destination path(str) for downloading
    sku_id_regex = re.compile('\\\\([0-9]*).xls')
    list_sku_id = []
    brands = []
    categories = []
    overall_sku_info = []
    config_file = read_config(CONFIG_PATH)

    for filepath in list_of_xlsx:

        sku_info = {}
        # Get SKU_ID and put inside sku_info dictionary
        sku_id = sku_id_regex.search(filepath).group(1)
        list_sku_id.append(sku_id)
        sku_info['sku_id'] = sku_id

        # Get brand and put inside sku_info_dictionary
        try:
            brand, category = read_xlsx_file(filepath)
            brands.append(brand)
            categories.append(category)
            sku_info['brand'] = brand
            sku_info['category'] = category

        except InvalidFileException as e:
            brand, category = read_xls_file(filepath)
            brands.append(brand)
            categories.append(category)
            sku_info['brand'] = brand
            sku_info['category'] = category

        # Use sort_file_path function to add file dest paths to each ID
        sorter = FileSort(sku_info)
        overall_sku_info = sorter.sort_file_path(config_file, overall_sku_info)
        download_dump_files = read_directory(PROJECT_PATH + config_file['defined_vars']['download_dump_folder'])

    # End of file meta reading to get list of SKUs to scrape from TaoSJ and their destination paths to download to
    # Ask user if they wish to download data from TaoSJ
    ask_to_download = input("Would you like to download data from TaoSJ? (YES/NO): ")

    if ask_to_download == "YES":
        # Login Process using selenium starts here
        driver = webdriver.Chrome('chromedriveedit.exe', options=option)
        driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
            "source": """Object.defineProperty(navigator, 'webdriver', {get: () => undefined})""",
        })
        driver.get(config_file['defined_vars']['login_url'])
        wait = WebDriverWait(driver, 10)
        login_process()

        time.sleep(10)

        # Navigate to '找宝贝' Tab
        shop_data()

        for count, sku in enumerate(overall_sku_info, start=1):
            taosj_sku_id = sku['sku_id']
            # Check if sku has already been downloaded
            if any(taosj_sku_id in s for s in download_dump_files):
                print(f'{taosj_sku_id} has already been downloaded. Will not download')
            else:
                print(f'{taosj_sku_id} has not been downloaded. Will download now.')
                print(f'Downloading file {count} out of {len_list_of_xlsx}')
                count = 0
                max_tries = 3
                while True:
                    try:
                        scrape_sku(taosj_sku_id)
                        break
                    except ElementClickInterceptedException as eleclickintercept:
                        print(eleclickintercept)
                        count += 1
                        if count == max_tries:
                            break
                        else:
                            continue
    else:
        print("Downloading of data from TaoSJ will not begin.")
    # Download of files to download-dump completed

    # Start shifting files to correct dest path
    ask_to_rename = input("Do you want to rename the files? (YES/NO): ")

    # Rename file type from xls to xlsx

    if ask_to_rename == 'YES':
        download_dump_files = read_directory(PROJECT_PATH + config_file['defined_vars']['download_dump_folder'])

        rename_files(download_dump_files)
        new_download_dump_files = read_directory_xlsx(PROJECT_PATH + config_file['defined_vars']['download_dump_folder'])

        sku_id_regex_download = re.compile('_([0-9]*)_')

        for files in new_download_dump_files:
            sku_id_download = sku_id_regex_download.search(files).group(1)
            shift_files(sku_id_download, files, overall_sku_info)
    else:
        print("Files have not been renamed and copied over to TaoSJ Data.")
        exit()

# TODO: Implement more try-catches to prevent errors

# TODO: Use logging module (Read up)

# TODO: Try and make file paths dynamic use os module